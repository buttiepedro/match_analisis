import io
import re
import unicodedata
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Annotated, Literal, Optional

import openpyxl
import pdfplumber
import xlrd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.permissions import Permission
from app.core.deps import get_current_user, require
from app.models import Event, MatchLineup, Player, Session, TimerState, Tournament, User, UserRole
from app.models.player import LineupStatus

router = APIRouter(prefix="/import")


# ─── PDF Parser helpers ───────────────────────────────────────────────────────


def _normalize_name(raw: str) -> str:
    parts = raw.split(",", 1)
    if len(parts) == 2:
        return f"{parts[1].strip()} {parts[0].strip()}".lower()
    return raw.strip().lower()


def _parse_lineup_table(rows: list) -> list[dict]:
    players = []
    header_found = False
    for row in rows:
        if not row:
            continue
        cells = [str(c or "").strip() for c in row]
        if not header_found:
            if cells and cells[0].lower() == "pos":
                header_found = True
            continue
        if all(not c for c in cells):
            continue
        try:
            pos = int(cells[0])
            dor = int(cells[1]) if len(cells) > 1 and cells[1] else pos
        except (ValueError, IndexError):
            continue
        nombre_raw = cells[2] if len(cells) > 2 else ""
        if not nombre_raw:
            continue
        raw_doc = cells[3] if len(cells) > 3 else ""
        doc_num = re.sub(r"\D", "", raw_doc)
        players.append({
            "pos": pos,
            "dor": dor,
            "nombre_pdf": nombre_raw,
            "nombre_norm": _normalize_name(nombre_raw),
            "status": "on_field" if pos <= 15 else "bench",
            "doc_num": doc_num,
        })
    return players


def _parse_incidencias(rows: list) -> list[dict]:
    events = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if not row:
            i += 1
            continue
        cells = [str(c or "").strip() for c in row]
        tiempo = cells[0] if cells else ""
        if tiempo not in ("1T", "2T"):
            i += 1
            continue
        try:
            minuto = int(cells[1])
        except (ValueError, IndexError):
            i += 1
            continue
        tipo = cells[2] if len(cells) > 2 else ""
        ptos_str = cells[3] if len(cells) > 3 else ""
        dorsal_str = cells[4] if len(cells) > 4 else ""
        obs = cells[5] if len(cells) > 5 else ""
        try:
            ptos = int(ptos_str) if ptos_str else 0
        except ValueError:
            ptos = 0
        try:
            dorsal = int(dorsal_str) if dorsal_str else None
        except ValueError:
            dorsal = None

        if tipo == "Try":
            converted = False
            if i + 1 < len(rows):
                nxt = rows[i + 1]
                if nxt and str(nxt[2] if len(nxt) > 2 else "").strip() == "Conversión":
                    converted = True
                    i += 1
            events.append({
                "tiempo": tiempo, "minuto": minuto, "tipo": "try",
                "dorsal": dorsal, "metadata": {"converted": converted},
            })
        elif tipo == "Penal" and ptos == 3:
            events.append({
                "tiempo": tiempo, "minuto": minuto, "tipo": "penalty",
                "dorsal": dorsal, "reason": "a_los_palos", "metadata": {"converted": True},
            })
        elif tipo == "Amarilla":
            events.append({
                "tiempo": tiempo, "minuto": minuto, "tipo": "yellow_card",
                "dorsal": dorsal, "observaciones": obs,
            })
        elif tipo == "Roja":
            events.append({
                "tiempo": tiempo, "minuto": minuto, "tipo": "red_card",
                "dorsal": dorsal,
            })
        elif tipo == "Se retiró":
            dorsal_in = None
            if i + 1 < len(rows):
                nxt = rows[i + 1]
                if nxt and str(nxt[2] if len(nxt) > 2 else "").strip() == "Ingresó":
                    try:
                        dorsal_in = int(str(nxt[4] if len(nxt) > 4 else "").strip())
                    except (ValueError, TypeError):
                        pass
                    i += 1
            if dorsal_in is not None:
                events.append({
                    "tiempo": tiempo, "minuto": minuto, "tipo": "substitution",
                    "dorsal_out": dorsal, "dorsal_in": dorsal_in,
                })
        i += 1
    return events


# ─── POST /import/lineup-pdf ──────────────────────────────────────────────────


@router.post("/lineup-pdf")
async def import_lineup_pdf(
    file: UploadFile = File(...),
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Se requiere un archivo PDF")

    content = await file.read()
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            if not pdf.pages:
                raise HTTPException(status_code=422, detail="PDF vacío")
            page = pdf.pages[0]
            text = page.extract_text() or ""
            tables = page.extract_tables()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el PDF: {e}")

    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]

    match_number = ""
    for line in lines:
        m = re.match(r"Tarjeta de partido\s+N[°o]?[:\s]*(\d+)", line, re.IGNORECASE)
        if m:
            match_number = m.group(1)
            break

    fecha_iso = None
    for line in lines:
        dm = re.search(r"(\d{4}-\d{2}-\d{2})", line)
        tm = re.search(r"\b(\d{2}:\d{2})\b", line)
        if dm and tm:
            fecha_iso = f"{dm.group(1)}T{tm.group(1)}:00"
            break

    local_team = visitante_team = ""
    local_score = visitante_score = 0
    skip_kws = {"Cancha", "Torneo", "División", "Tarjeta", "Pos", "Dor", "Tie.", "Instancia"}
    _date_time_re = re.compile(r"\d{4}-\d{2}-\d{2}|\b\d{2}:\d{2}\b")
    for line in lines:
        if any(kw in line for kw in skip_kws):
            continue
        # skip header info lines that contain date/time patterns (they can false-match)
        if _date_time_re.search(line):
            continue
        m = re.match(r"^([A-ZÁÉÍÓÚÑa-záéíóúñ\s]+?)\s+(\d+)\s+([A-ZÁÉÍÓÚÑa-záéíóúñ\s]+?)\s+(\d+)\s*$", line)
        if m:
            s1, s2 = int(m.group(2)), int(m.group(4))
            if 0 <= s1 <= 150 and 0 <= s2 <= 150:
                local_team = m.group(1).strip()
                local_score = s1
                visitante_team = m.group(3).strip()
                visitante_score = s2
                break

    lineup_tables: list = []
    incidencias_tables: list = []
    for tbl in (tables or []):
        if not tbl:
            continue
        first_cells = [str(c or "").strip() for c in tbl[0]]
        if "Pos" in first_cells and "Dor" in first_cells:
            lineup_tables.append(tbl)
        elif "Tie." in first_cells or "Incid." in first_cells:
            incidencias_tables.append(tbl)
        else:
            for row in tbl[1:4]:
                if row and any(str(c or "").strip() in ("1T", "2T") for c in row):
                    incidencias_tables.append(tbl)
                    break

    return {
        "match_number": match_number,
        "local_team": local_team,
        "visitante_team": visitante_team,
        "local_score": local_score,
        "visitante_score": visitante_score,
        "fecha": fecha_iso,
        "lineup_local": _parse_lineup_table(lineup_tables[0]) if lineup_tables else [],
        "lineup_visitante": _parse_lineup_table(lineup_tables[1]) if len(lineup_tables) > 1 else [],
        "incidencias_local": _parse_incidencias(incidencias_tables[0]) if incidencias_tables else [],
        "incidencias_visitante": _parse_incidencias(incidencias_tables[1]) if len(incidencias_tables) > 1 else [],
    }


# ─── POST /import/confirm ─────────────────────────────────────────────────────


class _LineupEntry(BaseModel):
    player_id: uuid.UUID
    jersey_number: int
    position: Optional[str] = None
    team: Literal["user", "rival"] = "user"
    status: Literal["on_field", "bench"] = "on_field"
    dni: Optional[str] = None


class _EventEntry(BaseModel):
    tiempo: str
    minuto: int
    tipo: str
    team: Literal["user", "rival"]
    reason: Optional[str] = None
    metadata: dict = {}


class ImportConfirmRequest(BaseModel):
    tournament_id: uuid.UUID
    home_team: str
    away_team: str
    scheduled_at: Optional[datetime] = None
    lineup: list[_LineupEntry]
    events: list[_EventEntry]


@router.post("/confirm")
async def import_confirm(
    body: ImportConfirmRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.partido_eventos))],
):
    tournament = await db.scalar(select(Tournament).where(Tournament.id == body.tournament_id))
    if not tournament:
        raise HTTPException(status_code=404, detail="Torneo no encontrado")
    if current_user.role != UserRole.superadmin and current_user.club_id != tournament.club_id:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    session = Session(
        id=uuid.uuid4(),
        tournament_id=tournament.id,
        home_team=body.home_team,
        away_team=body.away_team,
        scheduled_at=body.scheduled_at,
        half_duration_minutes=40,
        created_by=current_user.id,
    )
    db.add(session)
    await db.flush()

    db.add(TimerState(id=uuid.uuid4(), session_id=session.id))

    for entry in body.lineup:
        player = await db.scalar(
            select(Player).where(Player.id == entry.player_id, Player.is_active.is_(True))
        )
        if not player:
            raise HTTPException(status_code=404, detail=f"Jugador {entry.player_id} no encontrado")
        if entry.dni and not player.dni:
            player.dni = entry.dni
        db.add(MatchLineup(
            id=uuid.uuid4(),
            session_id=session.id,
            player_id=player.id,
            jersey_number=entry.jersey_number,
            position=entry.position,
            team=entry.team,
            status=LineupStatus(entry.status),
        ))

    now = datetime.now(timezone.utc)
    for ev in body.events:
        half = 1 if ev.tiempo == "1T" else 2
        db.add(Event(
            id=uuid.uuid4(),
            session_id=session.id,
            event_type=ev.tipo,
            half=half,
            timer_seconds=ev.minuto * 60,
            team=ev.team,
            reason=ev.reason,
            metadata_=ev.metadata,
            recorded_by=current_user.id,
            recorded_at=now,
        ))

    await db.commit()
    return {
        "session_id": str(session.id),
        "lineup_count": len(body.lineup),
        "event_count": len(body.events),
    }


# ─── Rugby positions catalogue ────────────────────────────────────────────────

RUGBY_POSITIONS: list[str] = [
    "01 - Pilar izquierdo",
    "02 - Hooker",
    "03 - Pilar derecho",
    "04 - Segundo línea",
    "05 - Segundo línea",
    "06 - Ala",
    "07 - Ala",
    "08 - Octavo",
    "09 - Medio scrum",
    "10 - Apertura",
    "11 - Wing izquierdo",
    "12 - Centro",
    "13 - Centro",
    "14 - Wing derecho",
    "15 - Full back",
]

# Map número (str) → posición estándar
_POS_BY_NUM: dict[str, str] = {p.split(" - ")[0].lstrip("0") or "0": p for p in RUGBY_POSITIONS}
_POS_BY_NUM.update({p.split(" - ")[0]: p for p in RUGBY_POSITIONS})  # con cero: "01", "02" …

# Map nombre normalizado → posición estándar (primera coincidencia gana)
_POS_BY_NAME: dict[str, str] = {}
for _p in RUGBY_POSITIONS:
    _key = unicodedata.normalize("NFKD", _p.split(" - ", 1)[1]).encode("ascii", "ignore").decode().lower()
    if _key not in _POS_BY_NAME:
        _POS_BY_NAME[_key] = _p


def _normalize_str(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower().strip()


def _map_position(raw: str) -> Optional[str]:
    """Convert any position string from the file to a standard RUGBY_POSITIONS entry."""
    raw = raw.strip()
    if not raw:
        return None
    # Format "03 - Pilar derecho" — extract leading number
    m = re.match(r"^0*(\d+)\s*[-–]?\s*(.*)", raw)
    if m:
        num = m.group(1)
        name_part = m.group(2).strip()
        # Try by number first
        candidate = _POS_BY_NUM.get(num) or _POS_BY_NUM.get(num.zfill(2))
        if candidate:
            return candidate
        # Fallback to name
        if name_part:
            norm = _normalize_str(name_part)
            return _POS_BY_NAME.get(norm)
    # Plain name
    return _POS_BY_NAME.get(_normalize_str(raw))


# ─── Excel reader ─────────────────────────────────────────────────────────────

# Column aliases → internal key
_COL_ALIASES: dict[str, str] = {
    # Las dos que agrega la exportación: identifican la fila, no describen al
    # jugador. Ver EXPORT_COLUMNS al final del módulo.
    "id": "id",
    "division": "division",
    "división": "division",
    "categoria": "division",
    "categoría": "division",
    "documento": "dni",
    "doc": "dni",
    "dni": "dni",
    "apellido": "last_name",
    "nombre": "first_name",
    "fecha nac.": "date_of_birth",
    "fecha nac": "date_of_birth",
    "fecha de nacimiento": "date_of_birth",
    "nacimiento": "date_of_birth",
    "sexo": "sex",
    "genero": "sex",
    "género": "sex",
    "o.social": "obra_social",
    "obra social": "obra_social",
    "obrasocial": "obra_social",
    "peso": "weight_kg",
    "estatura": "height_cm",
    "altura": "height_cm",
    "talla": "height_cm",
    "puesto": "position",
    "posicion": "position",
    "posición": "position",
    "email": "email",
    "correo": "email",
    "correo electronico": "email",
    "correo electrónico": "email",
    "tel.emergencia": "emergency_phone",
    "tel. emergencia": "emergency_phone",
    "telefono emergencia": "emergency_phone",
    "teléfono emergencia": "emergency_phone",
    "emergencia": "emergency_phone",
    "celular": "phone",
    "telefono": "phone",
    "teléfono": "phone",
    "movil": "phone",
    "móvil": "phone",
}


def _norm_col(s: str) -> str:
    return _normalize_str(s)


def _read_xlsx(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _read_xls(content: bytes) -> list[dict]:
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    return [ws.row_values(i) for i in range(ws.nrows)]


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # xlrd stores dates as floats — handle via xlrd if needed (already converted in _read_xls)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_weight(val) -> Optional[Decimal]:
    if val is None:
        return None
    s = re.sub(r"[^\d.,]", "", str(val)).replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_height(val) -> Optional[Decimal]:
    """Accepts 1.73, 1,73 (meters) or 173 (cm). Always returns cm."""
    if val is None:
        return None
    s = re.sub(r"[^\d.,]", "", str(val)).replace(",", ".")
    try:
        v = Decimal(s)
        # If value looks like meters (< 3), convert to cm
        if v < Decimal("3"):
            v = v * 100
        return v.quantize(Decimal("0.1"))
    except InvalidOperation:
        return None


def _parse_sex(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("M", "MASCULINO", "MALE", "H", "HOMBRE"):
        return "M"
    if s in ("F", "FEMENINO", "FEMALE", "MUJER"):
        return "F"
    return None


def _rows_to_dicts(raw_rows: list) -> list[dict]:
    """Map raw rows to list of normalised dicts using column aliases."""
    if not raw_rows:
        return []
    # Find header row (first row with at least 3 non-empty cells)
    header_idx = 0
    for i, row in enumerate(raw_rows[:5]):
        filled = sum(1 for c in row if c is not None and str(c).strip())
        if filled >= 3:
            header_idx = i
            break

    header = [_norm_col(str(c or "")) for c in raw_rows[header_idx]]
    col_map: dict[int, str] = {}
    for idx, h in enumerate(header):
        key = _COL_ALIASES.get(h)
        if key and key not in col_map.values():
            col_map[idx] = key

    records = []
    for row in raw_rows[header_idx + 1:]:
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        rec: dict = {}
        for idx, key in col_map.items():
            rec[key] = row[idx] if idx < len(row) else None
        records.append(rec)
    return records


# ─── POST /import/players-xlsx ────────────────────────────────────────────────

from app.models import Division  # noqa: E402 — avoids circular at module top


@router.post("/players-xlsx")
async def import_players_xlsx(
    file: Annotated[UploadFile, File(...)],
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.plantel_importar))],
    division_id: Annotated[Optional[uuid.UUID], Form()] = None,
):
    """
    Carga o actualiza jugadores desde una planilla.

    Acepta las dos formas de trabajar:

    - **Una división**: se manda `division_id` y todas las filas van ahí. Es la
      lista que le pasan al club a principio de año.
    - **El club entero**: la planilla trae una columna `División` por fila. Es lo
      que devuelve la exportación, para editar todo junto y volver a subirlo.

    Si están las dos, manda la de la fila: es más específica, y permite mover a
    alguien de división cambiando una celda.
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    division = None
    if division_id:
        division = await db.scalar(select(Division).where(Division.id == division_id))
        if not division:
            raise HTTPException(status_code=404, detail="División no encontrada")
        if current_user.role != UserRole.superadmin and current_user.club_id != division.club_id:
            raise HTTPException(status_code=403, detail="Acceso denegado")

    club_id = division.club_id if division else current_user.club_id
    if not club_id:
        raise HTTPException(status_code=400, detail="Indicá una división")

    # Divisiones del club por nombre normalizado, para resolver la columna
    # `División` de cada fila. "M17" y "m 17" tienen que caer en la misma.
    club_divisions = (
        await db.execute(
            select(Division).where(Division.club_id == club_id, Division.is_active.is_(True))
        )
    ).scalars().all()
    division_por_nombre = {_normalize_str(d.name): d for d in club_divisions}
    division_ids = [d.id for d in club_divisions]

    content = await file.read()
    try:
        raw_rows = _read_xlsx(content) if fname.endswith(".xlsx") else _read_xls(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    records = _rows_to_dicts(raw_rows)
    if not records:
        raise HTTPException(status_code=422, detail="El archivo no contiene filas de datos")

    created = updated = skipped = 0
    errors: list[dict] = []

    for row_num, rec in enumerate(records, start=2):
        # Compose name from apellido + nombre
        first = str(rec.get("first_name") or "").strip()
        last = str(rec.get("last_name") or "").strip()
        if first and last:
            name = f"{last} {first}"
        elif first:
            name = first
        elif last:
            name = last
        else:
            errors.append({"row": row_num, "reason": "Sin nombre ni apellido"})
            skipped += 1
            continue

        dni_raw = str(rec.get("dni") or "").strip()
        dni = re.sub(r"\D", "", dni_raw) or None

        pos_raw = str(rec.get("position") or "").strip()
        position = _map_position(pos_raw) if pos_raw else None

        dob = _parse_date(rec.get("date_of_birth"))
        sex = _parse_sex(rec.get("sex"))
        email = str(rec.get("email") or "").strip() or None
        phone = str(rec.get("phone") or "").strip() or None
        emergency = str(rec.get("emergency_phone") or "").strip() or None
        obra = str(rec.get("obra_social") or "").strip() or None
        weight = _parse_weight(rec.get("weight_kg"))
        height = _parse_height(rec.get("height_cm"))

        # A qué división va esta fila. La de la planilla gana sobre la del form.
        destino = division
        div_raw = str(rec.get("division") or "").strip()
        if div_raw:
            destino = division_por_nombre.get(_normalize_str(div_raw))
            if not destino:
                errors.append({"row": row_num, "reason": f"No existe la división '{div_raw}'"})
                skipped += 1
                continue
        if not destino:
            errors.append({"row": row_num, "reason": "Sin división: falta la columna o el campo"})
            skipped += 1
            continue

        # El ID manda sobre el DNI.
        #
        # Es lo que hace seguro exportar, editar y volver a subir: sin él, un
        # jugador sin DNI se duplica en cada vuelta, y corregirle un DNI mal
        # cargado crea uno nuevo en lugar de arreglar el que ya estaba — que es
        # justo lo que uno abre la planilla para hacer.
        existing: Optional[Player] = None
        id_raw = str(rec.get("id") or "").strip()
        if id_raw:
            try:
                existing = await db.scalar(
                    select(Player).where(
                        Player.id == uuid.UUID(id_raw),
                        Player.division_id.in_(division_ids),
                    )
                )
            except ValueError:
                errors.append({"row": row_num, "reason": f"ID inválido: '{id_raw}'"})
                skipped += 1
                continue
            if not existing:
                # Un ID que no existe casi siempre es una planilla de otro club, o
                # una fila copiada a mano. Crear un jugador suelto sería peor.
                errors.append({"row": row_num, "reason": "El ID no corresponde a este club"})
                skipped += 1
                continue

        if existing is None and dni:
            existing = await db.scalar(
                select(Player).where(
                    Player.dni == dni,
                    Player.division_id.in_(division_ids),
                    Player.is_active.is_(True),
                )
            )

        if existing:
            # Update existing player
            existing.name = name
            # Corregir el DNI es de los motivos principales para abrir la
            # planilla, y con el match por ID ya no hay riesgo de perder la fila.
            if dni:
                existing.dni = dni
            if position:
                existing.position = position
            if dob:
                existing.date_of_birth = dob
            if sex:
                existing.sex = sex
            if email:
                existing.email = email
            if phone:
                existing.phone = phone
            if emergency:
                existing.emergency_phone = emergency
            if obra:
                existing.obra_social = obra
            # Move to target division if different
            if existing.division_id != destino.id:
                existing.division_id = destino.id
            updated += 1
        else:
            player = Player(
                id=uuid.uuid4(),
                division_id=destino.id,
                name=name,
                position=position,
                dni=dni,
                date_of_birth=dob,
                sex=sex,
                email=email,
                phone=phone,
                emergency_phone=emergency,
                obra_social=obra,
            )
            db.add(player)
            created += 1

        # Peso y estatura del día, si vienen.
        if weight or height:
            from app.models.player import PlayerMeasurement

            player_id = existing.id if existing else player.id  # type: ignore[possibly-unbound]
            # Se pisa la medición de hoy en vez de agregar otra. Subir dos veces
            # la misma planilla —lo normal cuando se corrige una fila y se vuelve
            # a cargar— dejaba dos mediciones idénticas del mismo día, y la
            # evolución de peso pasaba a tener escalones que nadie midió.
            hoy = await db.scalar(
                select(PlayerMeasurement).where(
                    PlayerMeasurement.player_id == player_id,
                    PlayerMeasurement.measured_at == date.today(),
                )
            ) if existing else None

            if hoy:
                if weight:
                    hoy.weight_kg = weight
                if height:
                    hoy.height_cm = height
                hoy.recorded_by = current_user.id
            else:
                db.add(PlayerMeasurement(
                    player_id=player_id,
                    measured_at=date.today(),
                    weight_kg=weight,
                    height_cm=height,
                    recorded_by=current_user.id,
                ))

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar: {e}")

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(records),
    }


# ─── Exportar el plantel ──────────────────────────────────────────────────────
#
# El flujo que habilita: exportar, corregir treinta filas en Excel, volver a
# subir. Editar treinta jugadores de a uno en el celular no lo hace nadie.
#
# Para que la vuelta funcione, la planilla trae dos columnas que no son datos del
# jugador sino de identidad de la fila:
#
# - **ID**: el identificador interno. El importador lo mira **primero**. Sin él
#   el match es por DNI, y entonces un jugador sin DNI se duplica al volver, y
#   —peor— corregirle un DNI mal cargado crea un jugador nuevo en vez de
#   arreglar el que ya estaba, que es exactamente lo que uno va a querer hacer
#   con una planilla en la mano.
# - **División**: para poder exportar el club entero en un archivo y que cada
#   fila vuelva a la división que le corresponde. Cambiarla en la planilla mueve
#   al jugador, que es una forma cómoda de armar la pretemporada.
#
# El orden de las columnas es el de lectura de una planilla, no el del modelo:
# primero se busca a la persona, después se le corrigen los datos.

#: Encabezados de la exportación, en orden. Las claves son los nombres internos.
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("ID", "id"),
    ("DNI", "dni"),
    ("Apellido", "last_name"),
    ("Nombre", "first_name"),
    ("División", "division"),
    ("Puesto", "position"),
    ("Fecha nac.", "date_of_birth"),
    ("Sexo", "sex"),
    ("Email", "email"),
    ("Celular", "phone"),
    ("Tel. emergencia", "emergency_phone"),
    ("O.Social", "obra_social"),
]


def _split_name(name: str) -> tuple[str, str]:
    """
    Parte "Perez Juan" en apellido y nombre.

    Los jugadores se guardan con el nombre completo en un campo, pero la planilla
    los muestra en dos porque así vienen las listas de los clubes y así se
    ordenan. La primera palabra es el apellido: es como los arma el importador
    (`f"{apellido} {nombre}"`), así que la ida y la vuelta cierran.
    """
    partes = (name or "").strip().split()
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    return partes[0], " ".join(partes[1:])


@router.get("/players-xlsx")
async def export_players_xlsx(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.plantel_ver))],
    club_id: Annotated[Optional[uuid.UUID], Query()] = None,
    division_id: Annotated[Optional[uuid.UUID], Query()] = None,
):
    """
    Descarga el plantel como .xlsx, listo para editar y volver a subir.

    Sin parámetros exporta el club del usuario entero. Con `division_id`, sólo esa
    división.
    """
    from fastapi import Response

    from app.core.deps import get_division_or_404, scoped_division_ids

    if division_id:
        # Valida club y alcance por división de una.
        division = await get_division_or_404(division_id, db, current_user)
        divisiones = [division]
        nombre_archivo = f"plantel-{division.name}"
    else:
        target_club = club_id or current_user.club_id
        if not target_club:
            raise HTTPException(status_code=400, detail="Indicá un club o una división")
        if current_user.role != UserRole.superadmin and current_user.club_id != target_club:
            raise HTTPException(status_code=403, detail="Acceso denegado")

        query = select(Division).where(
            Division.club_id == target_club, Division.is_active.is_(True)
        )
        # Un entrenador con divisiones asignadas exporta las suyas, no el club
        # entero: el alcance por división vale igual acá que en la pantalla.
        permitidas = scoped_division_ids(current_user)
        if permitidas is not None:
            query = query.where(Division.id.in_(permitidas))
        divisiones = list((await db.execute(query.order_by(Division.name))).scalars().all())
        nombre_archivo = "plantel"

    if not divisiones:
        raise HTTPException(status_code=404, detail="No hay divisiones a las que llegues")

    nombres = {d.id: d.name for d in divisiones}
    players = (
        await db.execute(
            select(Player)
            .where(Player.division_id.in_(list(nombres)), Player.is_active.is_(True))
            .order_by(Player.division_id, Player.name)
        )
    ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantel"
    ws.append([encabezado for encabezado, _ in EXPORT_COLUMNS])

    for player in players:
        apellido, nombre = _split_name(player.name)
        ws.append([
            str(player.id),
            player.dni or "",
            apellido,
            nombre,
            nombres.get(player.division_id, ""),
            player.position or "",
            player.date_of_birth.isoformat() if player.date_of_birth else "",
            player.sex or "",
            player.email or "",
            player.phone or "",
            player.emergency_phone or "",
            player.obra_social or "",
        ])

    # Anchos a ojo: una planilla que abre con todas las columnas en 8 caracteres
    # obliga a acomodarla antes de poder leerla.
    for columna, ancho in zip("ABCDEFGHIJKL", (38, 12, 18, 18, 14, 16, 12, 6, 26, 16, 16, 20)):
        ws.column_dimensions[columna].width = ancho
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)

    # `Response` y no `StreamingResponse`: el archivo ya está entero en memoria,
    # así que no hay nada que ir mandando de a pedazos. Streamearlo además retrasa
    # el cierre de la sesión de base hasta que el cuerpo termina de salir, y una
    # conexión abierta de más por cada descarga no se paga por nada.
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nombre_archivo}.xlsx"',
            # Sin esto el navegador no ve el nombre del archivo en una respuesta
            # que atraviesa CORS.
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
