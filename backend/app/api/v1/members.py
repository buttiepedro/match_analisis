"""
Socios: padrón importable y estado de cuota.

La app **espeja** el estado que le da el sistema contable del club; no lo calcula
ni registra pagos.
"""
import io
import unicodedata
import uuid
from datetime import date, datetime, timezone
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import assert_club_access, get_club_or_404, get_current_user, require
from app.core.members import MassDeactivation, MemberRow, sync_members
from app.core.permissions import SOCIO, Permission
from app.core.roles import seed_club_roles
from app.core.security import get_password_hash
from app.models import Member, MemberImport, User, UserRole, user_roles
from app.schemas.member import (
    LinkableUser,
    MemberCreate,
    MemberImportResult,
    MemberImportLogEntry,
    MemberResponse,
    MemberUpdate,
    MyMembershipResponse,
)

router = APIRouter()


# ── Parser de Excel ───────────────────────────────────────────────────────────

def _normalize(text: Any) -> str:
    """
    Minúsculas, sin acentos y **sin puntuación**, para machear encabezados.

    La puntuación importa: un padrón real trae `N° Socio` y `Nro.`, y el `°` no es
    un acento —sobrevive a NFD— así que sin sacarlo la columna no matchea nunca.
    """
    s = str(text or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = "".join(c if c.isalnum() else " " for c in s)
    return " ".join(s.split())


#: Encabezados reconocidos por campo. El club exporta desde su sistema y nadie va
#: a renombrar columnas para que la app las lea.
HEADERS = {
    "document_id": {"dni", "documento", "nro documento", "n documento", "num documento"},
    "full_name": {"apellido y nombre", "nombre", "nombre y apellido", "socio"},
    "dues_up_to_date": {"al dia", "estado cuota", "estado", "cuota"},
    "category": {"categoria", "tipo socio", "tipo"},
    "member_number": {"n socio", "nro socio", "numero socio", "socio nro", "n de socio"},
    "email": {"email", "e mail", "correo"},
    "joined_on": {"fecha alta", "alta", "fecha de alta"},
}

#: Lo que el sistema contable puede escribir para "está al día".
TRUTHY = {"si", "sí", "s", "1", "true", "verdadero", "al dia", "ok", "x"}
FALSY = {"no", "n", "0", "false", "falso", "deudor", "debe", "moroso"}


def _parse_bool(raw: Any) -> Optional[bool]:
    value = _normalize(raw)
    if value in TRUTHY:
        return True
    if value in FALSY:
        return False
    return None


def _parse_date(raw: Any) -> Optional[date]:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(raw).strip(), fmt).date()
        except ValueError:
            continue
    return None


def parse_padron(content: bytes) -> tuple[list[MemberRow], list[dict]]:
    """
    Excel → filas normalizadas + errores por fila.

    Una fila mala **no descarta el import**: se importa el resto y se reporta qué
    quedó afuera con su número de fila.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise HTTPException(status_code=501, detail="openpyxl no está instalado")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo como Excel")

    sheet = wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    column_of: dict[str, int] = {}
    for index, cell in enumerate(header):
        name = _normalize(cell)
        for field, aliases in HEADERS.items():
            if name in aliases and field not in column_of:
                column_of[field] = index

    missing = [f for f in ("document_id", "full_name", "dues_up_to_date") if f not in column_of]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Faltan columnas obligatorias en el archivo: "
                + ", ".join({"document_id": "DNI", "full_name": "Nombre",
                             "dues_up_to_date": "Al día"}[f] for f in missing)
            ),
        )

    def cell(row: tuple, field: str) -> Any:
        index = column_of.get(field)
        return row[index] if index is not None and index < len(row) else None

    parsed: list[MemberRow] = []
    errors: list[dict] = []
    seen: set[str] = set()

    for number, row in enumerate(rows_iter, start=2):
        if row is None or all(c in (None, "") for c in row):
            continue

        document_id = str(cell(row, "document_id") or "").strip().replace(".", "")
        full_name = str(cell(row, "full_name") or "").strip()
        dues = _parse_bool(cell(row, "dues_up_to_date"))

        if not document_id:
            errors.append({"row": number, "reason": "Sin DNI"})
            continue
        if not full_name:
            errors.append({"row": number, "reason": "Sin nombre"})
            continue
        if dues is None:
            errors.append({"row": number, "reason": "Estado de cuota ilegible"})
            continue
        if document_id in seen:
            errors.append({"row": number, "reason": f"DNI {document_id} repetido en el archivo"})
            continue

        seen.add(document_id)
        parsed.append(
            MemberRow(
                document_id=document_id,
                full_name=full_name,
                dues_up_to_date=dues,
                category=str(cell(row, "category") or "").strip() or None,
                member_number=str(cell(row, "member_number") or "").strip() or None,
                email=str(cell(row, "email") or "").strip() or None,
                joined_on=_parse_date(cell(row, "joined_on")),
            )
        )

    return parsed, errors


# ── Import ────────────────────────────────────────────────────────────────────

@router.post("/clubs/{club_id}/members/import", response_model=MemberImportResult)
async def import_members(
    club_id: uuid.UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.socios_importar))],
    file: Annotated[UploadFile, File()],
    default_password: Annotated[str, Form()] = "",
    dry_run: Annotated[bool, Query()] = False,
    force: Annotated[bool, Query()] = False,
):
    """
    Sincroniza el padrón desde un Excel.

    Con `dry_run` devuelve qué haría **sin escribir**. Un import que daría de baja
    a más del 20% se rechaza salvo `force`: el error probable en una sincronización
    semanal es el archivo equivocado.
    """
    club = await get_club_or_404(club_id, db)
    assert_club_access(club, current_user)

    if not dry_run and len(default_password) < 8:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="La contraseña por defecto necesita al menos 8 caracteres",
        )

    rows, errors = parse_padron(await file.read())

    try:
        preview = await sync_members(
            club.id,
            rows,
            db,
            default_password=default_password,
            run_by=current_user.id,
            dry_run=dry_run,
            force=force,
        )
    except MassDeactivation as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"{exc}. Si el archivo es correcto, reintentá con force=true. "
                f"Se darían de baja: {', '.join(exc.preview.deactivated[:10])}"
            ),
        )

    preview.errors = errors

    if not dry_run:
        db.add(
            MemberImport(
                id=uuid.uuid4(),
                club_id=club.id,
                source="xlsx",
                created_count=len(preview.created),
                updated_count=len(preview.updated),
                deactivated_count=len(preview.deactivated),
                total_rows=preview.total_rows,
                errors=errors or None,
                run_by=current_user.id,
            )
        )
        await db.commit()

    return MemberImportResult(
        dry_run=dry_run,
        created=preview.created,
        updated=preview.updated,
        deactivated=preview.deactivated,
        total_rows=preview.total_rows,
        errors=errors,
    )


@router.get("/clubs/{club_id}/member-imports", response_model=list[MemberImportLogEntry])
async def list_member_imports(
    club_id: uuid.UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.socios_importar))],
):
    club = await get_club_or_404(club_id, db)
    assert_club_access(club, current_user)
    result = await db.execute(
        select(MemberImport)
        .where(MemberImport.club_id == club.id)
        .order_by(MemberImport.created_at.desc())
        .limit(30)
    )
    return result.scalars().all()


# ── Padrón ────────────────────────────────────────────────────────────────────

@router.get("/clubs/{club_id}/members", response_model=list[MemberResponse])
async def list_members(
    club_id: uuid.UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.socios_ver_todas))],
    search: Annotated[Optional[str], Query()] = None,
    only_debtors: Annotated[bool, Query()] = False,
):
    club = await get_club_or_404(club_id, db)
    assert_club_access(club, current_user)

    query = (
        select(Member)
        .where(Member.club_id == club.id, Member.is_active.is_(True))
        .options(selectinload(Member.user))
    )
    if only_debtors:
        query = query.where(Member.dues_up_to_date.is_(False))

    members = (await db.execute(query.order_by(Member.full_name))).scalars().all()

    if search:
        needle = _normalize(search)
        members = [
            m
            for m in members
            if needle in _normalize(m.full_name)
            or needle in (m.member_number or "")
            or needle in (m.user.document_id or "")
        ]

    return [
        MemberResponse(
            id=m.id,
            full_name=m.full_name,
            document_id=m.user.document_id,
            category=m.category,
            member_number=m.member_number,
            dues_up_to_date=m.dues_up_to_date,
            dues_synced_at=m.dues_synced_at,
        )
        for m in members
    ]


@router.get("/me/membership", response_model=MyMembershipResponse)
async def my_membership(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    """
    Estado propio del socio.

    Devuelve **siempre** `dues_synced_at`: sin la fecha, "estás al día" es un dato
    sin contexto que el socio puede tomar por actual cuando tiene tres semanas.
    """
    member = await db.scalar(select(Member).where(Member.user_id == current_user.id))
    if not member:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Este usuario no está vinculado a ningún socio",
        )
    return MyMembershipResponse(
        full_name=member.full_name,
        member_number=member.member_number,
        category=member.category,
        dues_up_to_date=member.dues_up_to_date,
        dues_synced_at=member.dues_synced_at,
        is_active=member.is_active,
    )


# ── Alta manual y asociación con un usuario ───────────────────────────────────
#
# El padrón es la fuente de verdad, pero no puede ser la **única** puerta. Un
# club que todavía no importó nada no tiene un solo socio, así que no hay forma
# de ver siquiera cómo se ve la pantalla. Y hay un caso permanente: el
# administrador del club también es socio, y su usuario ya existe.


def _to_member_response(member: Member, document_id: Optional[str]) -> MemberResponse:
    return MemberResponse(
        id=member.id,
        full_name=member.full_name,
        document_id=document_id,
        category=member.category,
        member_number=member.member_number,
        dues_up_to_date=member.dues_up_to_date,
        dues_synced_at=member.dues_synced_at,
    )


@router.get("/clubs/{club_id}/linkable-users", response_model=list[LinkableUser])
async def linkable_users(
    club_id: uuid.UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.socios_importar))],
):
    """Usuarios del club que todavía no son socios, para poder asociarlos."""
    club = await get_club_or_404(club_id, db)
    assert_club_access(club, current_user)

    ya_socios = select(Member.user_id).where(Member.club_id == club.id)
    users = (
        await db.execute(
            select(User)
            .where(
                User.club_id == club.id,
                User.is_active.is_(True),
                User.id.not_in(ya_socios),
            )
            .order_by(User.full_name)
        )
    ).scalars().all()

    return [LinkableUser.model_validate(u) for u in users]


@router.post(
    "/clubs/{club_id}/members",
    response_model=MemberResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_member(
    club_id: uuid.UUID,
    body: MemberCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.socios_importar))],
):
    """
    Da de alta un socio y lo asocia a un usuario.

    Tres caminos, y el orden importa:

    1. Con `user_id`, se asocia **ese** usuario. Es el administrador que además
       es socio.
    2. Sin `user_id`, se busca un usuario del club con ese DNI y se asocia.
       Asociar antes que crear evita el duplicado silencioso: dos cuentas para la
       misma persona, y la buena termina siendo la que no usa.
    3. Si no existe ninguno, se crea la cuenta con contraseña por defecto y
       cambio obligatorio en el primer ingreso, igual que en la importación.

    El DNI es obligatorio, y no por formalismo: la sincronización semanal matchea
    por DNI. Es lo único que hace que este socio sea **el mismo** que va a venir
    en el próximo export del contable — sin él, la primera importación lo daría
    de baja por ausente y crearía otro al lado.
    """
    club = await get_club_or_404(club_id, db)
    assert_club_access(club, current_user)

    document_id = "".join(ch for ch in (body.document_id or "") if ch.isdigit())
    if not document_id:
        raise HTTPException(status_code=400, detail="El DNI es obligatorio")

    account: Optional[User] = None

    if body.user_id:
        account = await db.scalar(
            select(User).where(User.id == body.user_id, User.club_id == club.id)
        )
        if not account:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Ese usuario no es de este club",
            )
        if account.document_id and account.document_id != document_id:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    f"El usuario ya tiene el DNI {account.document_id}. "
                    "Corregí uno de los dos antes de asociarlo."
                ),
            )
    else:
        account = await db.scalar(
            select(User).where(User.club_id == club.id, User.document_id == document_id)
        )

    if account:
        ya = await db.scalar(select(Member).where(Member.user_id == account.id))
        if ya:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"{ya.full_name} ya está cargado como socio",
            )

    full_name = (body.full_name or (account.full_name if account else "")).strip()
    if not full_name:
        raise HTTPException(status_code=400, detail="Falta el nombre")

    roles = await seed_club_roles(club.id, db)
    socio_role = roles.get(SOCIO)

    if account is None:
        if not body.default_password:
            raise HTTPException(
                status_code=400,
                detail=(
                    "No hay ningún usuario con ese DNI. Mandá una contraseña por "
                    "defecto para crearle la cuenta."
                ),
            )
        account = User(
            id=uuid.uuid4(),
            club_id=club.id,
            email=None,
            document_id=document_id,
            password_hash=get_password_hash(body.default_password),
            must_change_password=True,
            full_name=full_name,
            role=UserRole.player,  # placeholder del enum viejo; manda el rol Socio
        )
        db.add(account)
        await db.flush()
    elif not account.document_id:
        # Un usuario que entra con email no tiene DNI cargado. Se le pone, o el
        # padrón no lo va a reconocer y lo va a duplicar.
        account.document_id = document_id

    if socio_role:
        tiene = await db.scalar(
            select(user_roles.c.role_id).where(
                user_roles.c.user_id == account.id,
                user_roles.c.role_id == socio_role.id,
            )
        )
        if not tiene:
            await db.execute(
                user_roles.insert().values(user_id=account.id, role_id=socio_role.id)
            )

    member = Member(
        id=uuid.uuid4(),
        club_id=club.id,
        user_id=account.id,
        full_name=full_name,
        category=body.category,
        member_number=body.member_number,
        dues_up_to_date=body.dues_up_to_date,
        dues_synced_at=datetime.now(timezone.utc),
        is_active=True,
    )
    db.add(member)
    await db.commit()
    await db.refresh(member)

    return _to_member_response(member, document_id)


@router.patch("/clubs/{club_id}/members/{member_id}", response_model=MemberResponse)
async def update_member(
    club_id: uuid.UUID,
    member_id: uuid.UUID,
    body: MemberUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require(Permission.socios_importar))],
):
    """
    Corrige un socio a mano.

    Lo que se toque acá lo **pisa la próxima sincronización** si el padrón dice
    otra cosa: el sistema contable sigue siendo la fuente de verdad del estado de
    cuota. Sirve para el rato entre que alguien paga y llega el próximo export.
    """
    club = await get_club_or_404(club_id, db)
    assert_club_access(club, current_user)

    member = await db.scalar(
        select(Member)
        .where(Member.id == member_id, Member.club_id == club.id)
        .options(selectinload(Member.user))
    )
    if not member:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Socio no encontrado")

    if body.category is not None:
        member.category = body.category
    if body.member_number is not None:
        member.member_number = body.member_number
    if body.dues_up_to_date is not None:
        member.dues_up_to_date = body.dues_up_to_date
        # El socio ve "según el último dato del club, <fecha>". Cambiar el estado
        # sin mover la fecha haría que la pantalla mienta sobre su antigüedad.
        member.dues_synced_at = datetime.now(timezone.utc)
    if body.is_active is not None:
        member.is_active = body.is_active

    await db.commit()
    await db.refresh(member)
    return _to_member_response(member, member.user.document_id)
