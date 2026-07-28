"""
Exportar el plantel, editarlo en Excel y volver a subirlo.

El flujo que habilita: corregir treinta filas de una sentada. Editar treinta
jugadores de a uno en el celular no lo hace nadie.

Lo que se prueba acá es la **vuelta**, que es donde estaba el riesgo: sin una
columna de identidad, un jugador sin DNI se duplica en cada ciclo, y corregirle un
DNI mal cargado crea uno nuevo en vez de arreglar el que estaba — que es
exactamente para lo que uno abre la planilla.
"""
import io

import openpyxl
import pytest
from sqlalchemy import select

from app.models import Player, PlayerMeasurement

from tests.conftest import auth_header, login, make_user


@pytest.fixture
async def plantel(client, db, club_admin_ctx):
    """Tres jugadores en dos divisiones, uno de ellos sin DNI."""
    club = club_admin_ctx["club"]
    headers = club_admin_ctx["headers"]

    divisiones = {}
    for nombre in ("Primera", "M17"):
        res = await client.post(
            f"/clubs/{club.id}/divisions", json={"name": nombre}, headers=headers
        )
        assert res.status_code in (200, 201), res.text
        divisiones[nombre] = res.json()["id"]

    for div, nombre, dni in [
        ("Primera", "Perez Juan", "30111222"),
        ("Primera", "Gomez Sin Documento", None),
        ("M17", "Diaz Bruno", "30333444"),
    ]:
        body = {"name": nombre, "position": "Pilar"}
        if dni:
            body["dni"] = dni
        res = await client.post(
            f"/divisions/{divisiones[div]}/players", json=body, headers=headers
        )
        assert res.status_code in (200, 201), res.text

    return {"club": club, "headers": headers, "divisiones": divisiones}


async def exportar(client, ctx, **params):
    res = await client.get("/import/players-xlsx", params=params, headers=ctx["headers"])
    assert res.status_code == 200, res.text
    return openpyxl.load_workbook(io.BytesIO(res.content))


def filas(wb) -> list[dict]:
    ws = wb.active
    encabezados = [c.value for c in ws[1]]
    return [
        dict(zip(encabezados, [c.value for c in fila]))
        for fila in ws.iter_rows(min_row=2)
    ]


def a_bytes(wb) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def subir(client, ctx, wb, **data):
    return await client.post(
        "/import/players-xlsx",
        files={"file": ("plantel.xlsx", a_bytes(wb), "application/vnd.ms-excel")},
        data=data,
        headers=ctx["headers"],
    )


# ── Exportar ──────────────────────────────────────────────────────────────────

async def test_the_export_brings_the_whole_club_with_its_divisions(client, plantel):
    wb = await exportar(client, plantel)
    datos = filas(wb)

    assert len(datos) == 3
    assert {f["División"] for f in datos} == {"Primera", "M17"}
    assert {f["Apellido"] for f in datos} == {"Perez", "Gomez", "Diaz"}
    # El nombre completo se parte en apellido y nombre, como vienen las listas.
    perez = next(f for f in datos if f["Apellido"] == "Perez")
    assert perez["Nombre"] == "Juan"
    assert perez["DNI"] == "30111222"


async def test_one_division_can_be_exported_alone(client, plantel):
    wb = await exportar(client, plantel, division_id=plantel["divisiones"]["M17"])
    datos = filas(wb)
    assert len(datos) == 1
    assert datos[0]["Apellido"] == "Diaz"


async def test_every_row_carries_its_id(client, plantel):
    """Sin esta columna la vuelta no puede ser exacta."""
    datos = filas(await exportar(client, plantel))
    assert all(f["ID"] for f in datos)


# ── Volver a subir ────────────────────────────────────────────────────────────

async def test_uploading_the_export_unchanged_updates_and_creates_nothing(
    client, db, plantel
):
    """La prueba de fuego: un ciclo completo no puede inventar jugadores."""
    wb = await exportar(client, plantel)

    res = await subir(client, plantel, wb)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["created"] == 0, body
    assert body["updated"] == 3
    assert body["errors"] == []

    assert len((await db.execute(select(Player))).scalars().all()) == 3


async def test_a_player_without_dni_survives_the_round_trip(client, db, plantel):
    """Sin ID, éste se duplicaba en cada vuelta: no había con qué reconocerlo."""
    wb = await exportar(client, plantel)
    await subir(client, plantel, wb)
    await subir(client, plantel, wb)

    sin_dni = (
        await db.execute(select(Player).where(Player.dni.is_(None)))
    ).scalars().all()
    assert len(sin_dni) == 1


async def test_fixing_a_wrong_dni_edits_the_player_instead_of_creating_one(
    client, db, plantel
):
    """
    El motivo principal para abrir la planilla.

    Con match por DNI, corregirlo creaba un jugador nuevo y dejaba al viejo con el
    documento mal — el problema que uno venía a arreglar, ahora duplicado.
    """
    wb = await exportar(client, plantel)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2):
        if fila[2].value == "Perez":
            fila[1].value = "99999999"

    res = await subir(client, plantel, wb)
    assert res.json()["created"] == 0, res.json()

    perez = await db.scalar(select(Player).where(Player.name == "Perez Juan"))
    assert perez.dni == "99999999"
    assert len((await db.execute(select(Player))).scalars().all()) == 3


async def test_changing_the_division_column_moves_the_player(client, db, plantel):
    """Una forma cómoda de armar la pretemporada: mover treinta filas de una."""
    wb = await exportar(client, plantel)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2):
        if fila[2].value == "Diaz":
            fila[4].value = "Primera"

    res = await subir(client, plantel, wb)
    assert res.status_code == 200, res.text

    datos = filas(await exportar(client, plantel))
    diaz = next(f for f in datos if f["Apellido"] == "Diaz")
    assert diaz["División"] == "Primera"


async def test_an_edited_field_lands(client, db, plantel):
    wb = await exportar(client, plantel)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2):
        if fila[2].value == "Gomez":
            fila[8].value = "gomez@club.com"

    await subir(client, plantel, wb)

    gomez = await db.scalar(select(Player).where(Player.name.like("Gomez%")))
    assert gomez.email == "gomez@club.com"


# ── Lo que no se acepta ───────────────────────────────────────────────────────

async def test_a_row_with_an_unknown_division_is_reported_not_guessed(client, plantel):
    wb = await exportar(client, plantel)
    ws = wb.active
    ws.cell(row=2, column=5).value = "Intermedia"

    res = await subir(client, plantel, wb)
    assert res.status_code == 200, res.text
    assert res.json()["skipped"] == 1
    assert "Intermedia" in res.json()["errors"][0]["reason"]


async def test_an_id_from_another_club_is_refused(client, db, plantel):
    """Una planilla del club de al lado no puede sembrar jugadores acá."""
    import uuid as _uuid

    wb = await exportar(client, plantel)
    ws = wb.active
    ws.cell(row=2, column=1).value = str(_uuid.uuid4())

    res = await subir(client, plantel, wb)
    assert res.json()["skipped"] == 1
    assert "ID" in res.json()["errors"][0]["reason"]
    assert len((await db.execute(select(Player))).scalars().all()) == 3


# ── Mediciones ────────────────────────────────────────────────────────────────

async def test_reuploading_weights_does_not_stack_measurements(client, db, plantel):
    """
    Subir dos veces la misma planilla dejaba dos mediciones idénticas del mismo
    día, y la evolución de peso pasaba a tener escalones que nadie midió.
    """
    wb = await exportar(client, plantel)
    ws = wb.active
    ws.cell(row=1, column=13).value = "Peso"
    for fila in range(2, 5):
        ws.cell(row=fila, column=13).value = 90

    await subir(client, plantel, wb)
    await subir(client, plantel, wb)

    mediciones = (await db.execute(select(PlayerMeasurement))).scalars().all()
    assert len(mediciones) == 3, "una por jugador, no dos"


# ── Permisos ──────────────────────────────────────────────────────────────────

async def test_exporting_requires_seeing_the_squad(client, db, plantel):
    from app.models import UserRole

    jugador = await make_user(
        db, email="jug@example.com", role=UserRole.player, club_id=plantel["club"].id
    )
    tokens = await login(client, jugador.email)

    res = await client.get(
        "/import/players-xlsx", headers=auth_header(tokens["access_token"])
    )
    assert res.status_code == 403
